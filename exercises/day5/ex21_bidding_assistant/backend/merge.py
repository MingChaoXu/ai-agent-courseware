# -*- coding: utf-8 -*-
"""
Excel 文件合并 - 将多个 phase2 Excel 文件合并为一个。
支持两种合并方式：
  1. 按文件选择：指定要合并的文件名列表
  2. 按时间范围：指定起止日期，合并该时间段内生成的文件

合并时按 (标题, 发布日期) 去重，保留首次出现。

注意：
  读取/写入都按"实际表头"自适应列数，兼容：
    - 老版本 10 列（无 客户经理/bu名称/方案跟进情况）
    - 新版本 13 列（含上述三列）
  合并结果的列顺序 = 标准 13 列 + 任何"老文件中独有"的新列（理论上没有）。
"""
import os
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PHASE2_DIR = os.path.join(BASE_DIR, "data", "phase2")

# 标准列定义（与 pipeline_v3._COLUMNS 一致）
# merge.py 不再硬编码 10/13，而是按"源文件实际表头"自适应
STANDARD_COLUMNS = [
    ("title",      "公告标题",     45),
    ("tenderer",   "招标人",       20),
    ("date",       "发布日期",     13),
    ("result_due", "开标日期",     13),
    ("amount",     "预算金额",     14),
    ("is_it",      "信息化",        8),
    ("public_due", "公告期限",     14),
    ("source",     "数据来源",     10),
    ("matched",    "是否所属客户", 20),
    ("manager",    "客户经理",     14),
    ("bu",         "bu名称",       22),
    ("followup",   "方案跟进情况", 16),
    ("url",        "公告链接",     50),
]

# 中文表头 → 字段 key 的反查表（兼容老文件的中文别称）
_HEADER_TO_KEY = {header: key for key, header, _ in STANDARD_COLUMNS}


def _normalize_date(date_str: str) -> str:
    match = re.search(r'(\d{4}-\d{2}-\d{2})', str(date_str) or "")
    return match.group(1) if match else str(date_str or "").strip()


def _detect_columns(header_row: tuple) -> tuple[list, dict, dict]:
    """根据源文件表头推断列结构。

    返回:
        columns   : [(key, 中文表头, 默认宽度), ...]  按源文件实际列顺序
        header_to_key : {中文表头: key}
        key_to_index : {key: 在源文件中的列索引}
    """
    if not header_row:
        return [], {}, {}
    columns = []
    header_to_key = {}
    key_to_index = {}
    # 缺省宽度 12（中文字段）
    default_width = 12
    for idx, name in enumerate(header_row):
        name = str(name).strip() if name is not None else ""
        if not name:
            continue
        # 查表 → 字段 key
        key = _HEADER_TO_KEY.get(name)
        if key is None:
            # 未知列：用字段名作为 key（保留原始数据，不丢字段）
            key = f"extra_{idx}"
            columns.append((key, name, default_width))
        else:
            width = next((w for k, h, w in STANDARD_COLUMNS if k == key), default_width)
            columns.append((key, name, width))
        header_to_key[name] = key
        key_to_index[key] = idx
    return columns, header_to_key, key_to_index


def _read_rows(xlsx_path: str) -> tuple[list, list]:
    """读取 Excel 文件的所有数据行，返回 (rows, columns)。

    rows  : list[dict]，每个 dict 至少含所有 STANDARD_COLUMNS 中的 key，
            缺失列值为空字符串
    columns: [(key, 表头, 宽度)]，按源文件实际表头顺序

    自适应兼容：
        - 10 列老格式（无 manager/bu/followup）
        - 13 列新格式
        - 其他任意列（不会丢字段，统一塞到标准字段后面）
    """
    if not os.path.exists(xlsx_path):
        return [], list(STANDARD_COLUMNS)
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return [], list(STANDARD_COLUMNS)

        # 1) 按表头检测列结构
        detected_cols, _, _ = _detect_columns(header_row)

        # 2) 合并列：标准列优先 + 老文件独有列（理论上没有，但保留鲁棒性）
        #    按"标准列顺序"输出，未知列追加在 url 之后
        standard_keys = [k for k, _, _ in STANDARD_COLUMNS]
        standard_set = set(standard_keys)
        ordered_cols = list(STANDARD_COLUMNS)
        for col in detected_cols:
            if col[0] not in standard_set:
                ordered_cols.append(col)

        # 3) 字段 key → 在源文件中的列索引
        name_to_idx = {name: idx for idx, name in enumerate(header_row) if name is not None}

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            record = {}
            for key, header, _ in ordered_cols:
                idx = name_to_idx.get(header)
                record[key] = row[idx] if idx is not None and idx < len(row) else ""
            rows.append(record)
        wb.close()
        return rows, ordered_cols
    except Exception as e:
        print(f"[merge] 读取失败 {xlsx_path}: {e}")
        return [], list(STANDARD_COLUMNS)


def _get_files_by_time(start_date: str, end_date: str) -> list:
    """按文件修改时间筛选 phase2 目录下的 Excel 文件"""
    if not os.path.exists(PHASE2_DIR):
        return []
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        # end_date 包含当天，所以加一天
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
    except ValueError:
        return []

    files = []
    for name in os.listdir(PHASE2_DIR):
        if not name.lower().endswith(".xlsx"):
            continue
        full = os.path.join(PHASE2_DIR, name)
        if not os.path.isfile(full):
            continue
        mtime = datetime.fromtimestamp(os.path.getmtime(full))
        if start <= mtime < end:
            files.append(name)
    files.sort()
    return files


def merge_files(file_names: list, output_name: str = None) -> str:
    """
    合并指定的 phase2 Excel 文件。
    file_names: 文件名列表（位于 PHASE2_DIR 下）
    output_name: 输出文件名（不含路径），不指定则自动生成
    返回：输出文件路径
    """
    all_rows = []
    merged_cols = list(STANDARD_COLUMNS)  # 收集所有源文件中出现过的列
    seen_col_set = set(k for k, _, _ in merged_cols)

    for name in file_names:
        path = os.path.join(PHASE2_DIR, name) if not os.path.isabs(name) else name
        rows, cols = _read_rows(path)
        all_rows.extend(rows)
        print(f"[merge] {name}: 读取 {len(rows)} 行 (列数={len(cols)})")
        # 收集"额外列"（老文件若有未在标准列里的字段，保留下来）
        for c in cols:
            if c[0] not in seen_col_set:
                merged_cols.append(c)
                seen_col_set.add(c[0])

    # 去重
    seen = set()
    unique = []
    for r in all_rows:
        key = (str(r.get("title", "")).strip(), _normalize_date(r.get("date", "")))
        if key not in seen:
            seen.add(key)
            unique.append(r)

    print(f"[merge] 合并前 {len(all_rows)} 行 → 去重后 {len(unique)} 行")
    print(f"[merge] 合并列数 = {len(merged_cols)}（标准 13 列 + 额外 {len(merged_cols) - len(STANDARD_COLUMNS)} 列）")

    # 生成输出文件
    if output_name is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"merged_{ts}.xlsx"
    output_path = os.path.join(PHASE2_DIR, output_name)

    _write_excel(unique, output_path, merged_cols)
    print(f"[merge] 合并完成 → {output_path}")
    return output_path


def merge_by_time(start_date: str, end_date: str, output_name: str = None) -> str:
    """
    按时间范围合并 phase2 Excel 文件。
    start_date / end_date: "YYYY-MM-DD" 格式
    返回：输出文件路径
    """
    files = _get_files_by_time(start_date, end_date)
    if not files:
        print(f"[merge] 时间范围 {start_date} ~ {end_date} 内无文件")
        return None
    print(f"[merge] 时间范围内找到 {len(files)} 个文件: {files}")
    return merge_files(files, output_name)


def _write_excel(records: list, xlsx_path: str,
                 columns: list = None) -> None:
    """将记录写入带格式的 Excel 文件。

    columns: 自定义列定义 [(key, 表头, 宽度), ...]，缺省用标准 13 列
    """
    if columns is None:
        columns = STANDARD_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "招标公告"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    link_font = Font(color="0563C1", underline="single", size=9)
    normal_font = Font(size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="BFBFBF")
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    # 表头
    headers = [col[1] for col in columns]
    ws.append(headers)
    for col_idx, (_, header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 找到 url 列（若有）
    url_col_idx = None
    for i, (k, _, __) in enumerate(columns):
        if k == "url":
            url_col_idx = i + 1
            break

    for row_idx, record in enumerate(records, start=2):
        for col_idx, (key, _, __) in enumerate(columns, start=1):
            value = record.get(key) or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = left_align
            cell.font = normal_font
            if col_idx == url_col_idx and value:
                cell.hyperlink = value
                cell.font = link_font
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="EBF3FB")
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, len(records) + 2):
        ws.row_dimensions[row_idx].height = 32

    wb.save(xlsx_path)


# 保留向后兼容的别名（避免其他模块引用时报错）
COLUMNS = STANDARD_COLUMNS
