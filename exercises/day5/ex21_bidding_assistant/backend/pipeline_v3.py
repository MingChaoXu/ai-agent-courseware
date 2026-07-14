"""
pipeline_v3.0.py
招标公告 + 采购意向 全流程分阶段运行：采集 → 去重 → AI提取字段 → 客户匹配 → 保存

流程分为两个阶段：
  Phase 1: 采集各个网站的招标信息
      - get_gov    : 苏州市政府采购平台「招标公告」（Selenium）
      - get_gov2   : 苏州市政府采购平台「采购意向」（Selenium + requests）
      - get_jc     : 金采平台公告
      - get_public : 公共资源平台公告
      - get_js     : 江苏省政府采购平台「招标公告」（公开招标）
      - get_js2    : 江苏省政府采购平台「采购意向」
      每个 get_*.py 生成带时间戳的中间文件（intermediate_<source>_YYYYMMDD_HHMMSS.json），
      合并后自动删除中间文件。
      最终输出：./data/phase1/phase1_collected_YYYYMMDD_HHMMSS.json

  Phase 2: 处理采集到的信息，提取字段并导出
      - 读取 Phase 1 输出的 JSON 文件
      - 调用 SiliconFlow 模型逐条提取 amount / is_it / public_due / result_due / tenderer
      - 自动匹配客户（如果存在 kehu-0619.xlsx）
      - 只生成 Excel（不生成 JSON）
      最终输出：./data/phase2/phase2_extracted_YYYYMMDD_HHMMSS.xlsx

运行方式：
  # 运行完整流程（Phase 1 + Phase 2）
  python pipeline_v3.py

  # 仅运行 Phase 1（采集）
  python pipeline_v3.py --phase 1

  # 仅运行 Phase 2（处理），需指定输入文件
  python pipeline_v3.py --phase 2 --input data/phase1/phase1_collected_20260414_152030.json

注意：
  - 如需客户匹配功能，请在同一目录下放置 kehu-0619.xlsx 文件
  - kehu-0619.xlsx 含三列：客户名称(grid_name_t0)、客户经理(area_name_t0)、bu名称(area_name_zj_t0)
  - get_gov2 采集的是「采购意向」公告，source 标签为 "gov2"
  - get_js  采集的是江苏省级「招标公告」（公开招标），source 标签为 "js"
  - get_js2 采集的是江苏省级「采购意向」，source 标签为 "js2"
"""

import os
import sys
import json
import re
import time
import traceback
import argparse
from datetime import datetime
from difflib import SequenceMatcher

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from collectors import get_gov, get_gov2, get_jc, get_public, get_js, get_js2
import process
import config as _config
import history as _history


# ================================================================== #
#  路径配置
# ================================================================== #
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# ================================================================== #
#  客户匹配配置
# ================================================================== #
KEHU_XLSX = os.path.join(BASE_DIR, "kehu-0619.xlsx")    # 客户清单文件（含客户经理/bu名称）


def _get_threshold():
    """从 config.py 读取客户匹配阈值"""
    return _config.get_threshold()


# ================================================================== #
#  输出目录配置（统一持久化保存）
# ================================================================== #
PHASE1_DIR = os.path.join(BASE_DIR, "data", "phase1")   # Phase 1 最终文件目录
PHASE2_DIR = os.path.join(BASE_DIR, "data", "phase2")   # Phase 2 最终文件目录


# ================================================================== #
#  数据来源列表（含 gov2 采购意向 + js2 江苏采购意向）
# ================================================================== #

SOURCES = [
    ("gov",    get_gov.run),
    ("gov2",   get_gov2.run),
    ("jc",     get_jc.run),
    ("public", get_public.run),
    ("js",     get_js.run),
    ("js2",    get_js2.run),
]


# ================================================================== #
#  辅助函数：日期标准化、去重
# ================================================================== #

def normalize_date(date_str: str) -> str:
    """统一日期格式为 YYYY-MM-DD"""
    match = re.search(r'(\d{4}-\d{2}-\d{2})', date_str or "")
    return match.group(1) if match else date_str


def dedup(records: list) -> list:
    """按 (title, date) 去重，保留首次出现"""
    seen = set()
    unique = []
    for item in records:
        key = (item.get("title", "").strip(), normalize_date(item.get("date", "")))
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


# ================================================================== #
#  Phase 1: 采集各个网站的招标信息
# ================================================================== #

def phase1_collect() -> tuple:
    """
    从六个来源采集公告，返回 (合并去重后的列表, 中间文件路径列表)。
    任一来源失败不影响其他来源继续运行。
    每个 get_*.py 会生成带时间戳的中间文件，由调用方在使用后统一删除。
    """
    all_data = []
    intermediate_files = []
    summary = {}

    for name, func in SOURCES:
        print(f"\n{'='*50}")
        print(f"  开始采集: [{name}]  {datetime.now().strftime('%H:%M:%S')}")
        print(f"{'='*50}")
        try:
            data, intermediate_file = func()
            if intermediate_file:
                intermediate_files.append(intermediate_file)
            for item in data:
                item["date"]   = normalize_date(item.get("date", ""))
                item["source"] = name
            all_data.extend(data)
            summary[name] = {"status": "ok", "count": len(data)}
            print(f"[{name}] 采集完成，获取 {len(data)} 条")
        except Exception as e:
            summary[name] = {"status": "error", "error": str(e)}
            print(f"[{name}] 采集失败: {e}")
            traceback.print_exc()

    try:
        result = dedup(all_data)

        print(f"\n{'='*50}")
        print(f"  采集汇总")
        for name, info in summary.items():
            if info["status"] == "ok":
                print(f"  [{name}]  采集  {info['count']} 条")
            else:
                print(f"  [{name}]  失败  {info['error']}")
        print(f"\n  合计 {len(all_data)} 条 → 去重后 {len(result)} 条")
        print(f"{'='*50}")

        return result, intermediate_files
    except Exception as e:
        # 去重或汇总阶段异常：仍返回已采集的数据和中间文件，避免泄漏
        print(f"[phase1_collect] 去重/汇总阶段异常: {e}")
        traceback.print_exc()
        return all_data, intermediate_files


def cleanup_intermediate_files(file_paths: list) -> None:
    """删除 get_*.py 生成的中间文件"""
    print(f"\n清理中间文件（共 {len(file_paths)} 个）...")
    for fp in file_paths:
        if fp and os.path.exists(fp):
            try:
                os.remove(fp)
                print(f"  已删除: {fp}")
            except Exception as e:
                print(f"  删除失败 {fp}: {e}")
    print("中间文件清理完成。")


def run_phase1(output_file: str = None) -> tuple[list, str]:
    """
    运行 Phase 1，保存采集结果到 ./data/phase1/ 目录下（带时间戳）。
    每个 get_*.py 生成的中间文件在合并后自动删除。
    即使采集过程异常，也会在 finally 中清理中间文件，避免残留。
    返回：(数据列表, 输出文件路径)
    """
    print("\n" + "█"*50)
    print("  Phase 1: 采集各个网站的招标信息")
    print("█"*50)

    # 确保输出目录存在
    os.makedirs(PHASE1_DIR, exist_ok=True)

    intermediate_files = []
    try:
        collected, intermediate_files = phase1_collect()

        if not collected:
            print("\n没有采集到任何公告，Phase 1 结束。")
            return [], ""

        # 统一命名：phase1_collected_YYYYMMDD_HHMMSS.json，保存到 ./data/phase1/
        if output_file is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(PHASE1_DIR, f"phase1_collected_{ts}.json")

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(collected, f, ensure_ascii=False, indent=2)

        print(f"\nPhase 1 完成！采集结果已保存至 {output_file}（共 {len(collected)} 条）")

        return collected, output_file
    finally:
        # 无论成功、空结果还是异常，都清理中间文件
        cleanup_intermediate_files(intermediate_files)


# ================================================================== #
#  Phase 2: AI 字段提取
# ================================================================== #

def extract_fields(records: list) -> list:
    """
    对采集到的公告列表逐条调用 AI 提取字段。
    返回包含完整字段的结果列表（提取失败的字段填 None）。
    当 AI 熔断器开启时，剩余记录直接填 None 跳过，避免长时间无效等待。
    """
    results = []
    total = len(records)
    breaker_opened = False
    for idx, item in enumerate(records, 1):
        # 熔断器已开启：剩余记录直接跳过
        if breaker_opened or process.ai_breaker_state() == "open":
            if not breaker_opened:
                print("  [熔断器] AI 接口已熔断，剩余记录跳过字段提取。")
                breaker_opened = True
            row = {
                "title":      item.get("title"),
                "date":       item.get("date"),
                "url":        item.get("url"),
                "source":     item.get("source"),
                "content":    item.get("content"),
                "amount":     None, "is_it": None, "public_due": None,
                "result_due": None, "tenderer": None, "matched": None,
            }
            results.append(row)
            continue
        print(f"  [{idx}/{total}] 提取字段: {item['title'][:40]}...")
        extracted = process.extract_info_with_siliconflow(item)
        row = {
            "title":      item.get("title"),
            "date":       item.get("date"),
            "url":        item.get("url"),
            "source":     item.get("source"),
            "content":    item.get("content"),
            "amount":     extracted["amount"]     if extracted else None,
            "is_it":      extracted["is_it"]      if extracted else None,
            "public_due": extracted["public_due"] if extracted else None,
            "result_due": extracted["result_due"] if extracted else None,
            "tenderer":   extracted["tenderer"]   if extracted else None,
            "matched":    None,  # 客户匹配结果，在导出 Excel 时填充
        }
        results.append(row)
        time.sleep(1)   # 避免触发频率限制
    return results


# ================================================================== #
#  第三阶段：客户匹配
# ================================================================== #

def preprocess(text: str) -> str:
    """统一小写，剔除空格、括号及其内容（减少干扰）。
    增强：去地址后缀、统一机构简称（委/局/院/会/校 等）。

    改动说明：
      - 不再剥离「院/医院/学校/学院」等关键法人后缀，避免
        "苏州市立医院" 被剥成 "苏州市立"，导致与无关机构误匹配。
      - 仅抹平「委员会→委」「有限公司→公司」等明确是简称/全称差异的后缀。
    """
    if not text:
        return ""
    text = text.lower().strip()
    text = re.sub(r'\s+', '', text)
    # 抹平全/半角空格
    text = text.replace('\u3000', '')
    # 去掉括号及其内容（与之前一致）
    text = re.sub(r'（.*?）', '', text)
    text = re.sub(r'\(.*?\)', '', text)
    # 去掉地址/分院后缀（"苏州大学独墅湖校区" → "苏州大学"）
    # 兼容：括号在前面已处理，剩余的「校区/院区/分部/基地/本部/总院/分院/东/西/新/老校区」
    text = re.sub(
        r'(独墅湖|天赐庄|阳澄湖|本部|总院|分院|东校区|西校区|新校区|老校区)'
        r'(校区|院区|分部|基地|医院|本部|院|校|中心)?$',
        '', text
    )
    # 抹平 委会/委/局/所/校/院 等单字差异（"发展和改革委员会" → "发展和改革委"）
    # 目的：让 "苏州市发展和改革委员会" 与 "苏州市发展和改革委" 都能匹配
    text = re.sub(r'委员会$', '委', text)
    text = re.sub(r'管理局$', '局', text)
    text = re.sub(r'事务所$', '所', text)
    # 公司类后缀：仅抹平"有限责任公司/股份公司/有限公司→公司"这类明确简称差异
    text = re.sub(r'有限责任公司$', '', text)
    text = re.sub(r'有限公司$', '', text)
    text = re.sub(r'股份公司$', '', text)
    text = re.sub(r'公司$', '', text)
    return text


# 政府/事业单位常见的「停用词」集合：仅当去除后还能识别出客户才考虑
# 这里用于 "客户名是 tenderer 子串" 的反向匹配
_GOV_SUFFIXES = (
    "人民政府", "人民政府办公室", "人民政府办公室", "管理委员会",
    "管委会", "管理委员会办公室",
)


# ================================================================== #
# 上下级/分支机构的特征词：用于识别「不同主体」的误匹配
#  仅收录"业务/差异化/子机构"词，机构类型词（医院/学校/中心）不收录
# ================================================================== #
_BRANCH_TOKENS = (
    # 上下级关系：公安局/交警支队/派出所/分局
    "交通警察", "交警", "派出所", "分局", "支队", "大队",
    # 分支机构：分行/支行/营业部
    "分行", "支行", "营业", "营业部", "业务部", "办事处", "代表处",
    # 附属/附属医院
    "附属", "附属医院",
    # 子公司层级词（注意"集团"放这里因为子公司/母公司结构常见）
    "控股", "集团", "集团有限", "集团有限公司",
    # 校区/分部
    "校区", "分部", "院区", "本部",
    # 业务/法人主体差异词（不同子业务/不同子公司视为不同主体）
    "印刷", "演艺", "培训", "旅游", "风景",
    "东方水城", "水上", "文创", "演艺", "传媒",
    "租赁", "置业", "经营", "经营管理",
    "检测", "监测", "体检", "医疗", "养老", "康养",
    "投资", "发展", "管理", "管理公司", "管理有限",
)


# ================================================================== #
#  主体判别词：若 tenderer 与候选各含其一（且不重叠）→ 视为不同主体
#  用于识别「附属第一医院 vs 附属第二医院」「虎丘山 vs 盘门」
#  「江苏苏州 vs 江苏东吴」等结构相同但关键标识不同的误匹配
# ================================================================== #
_DIFFERENTIATOR_TOKENS = (
    # 序数词（医院/学校常见）
    "第一", "第二", "第三", "第四", "第五",
    "第六", "第七", "第八", "第九", "第十",
    # 数字（医院"附属1院/附属2院"）
    "1院", "2院", "3院", "4院", "5院",
    "一分院", "二分院", "三分院",
    # 单字序数（preprocess 后会剥离"第"）—— 仅在 diff 片段中触发
    # 实际匹配逻辑见 _has_different_differentiator 内对单字的处理
    # 苏州市各区/县级区划
    "姑苏区", "工业园区", "园区", "吴中区", "相城区", "虎丘区", "高新区", "新区",
    # 苏州著名地名/景区（风景名胜区管理处 类）
    "虎丘山", "盘门", "金鸡湖", "独墅湖", "阳澄湖", "石湖",
    # 银行/企业字号（江苏苏州/江苏东吴）
    "东吴", "吴中", "姑苏",
    # 集团 vs 区域子公司（吴中产业投资 vs 产业投资）
    "吴中", "相城", "姑苏", "园区",
)

# 中文单字序数（用于 diff 片段判断）—— 短至 1 字符，必须谨慎
_CHINESE_DIGITS = ("一", "二", "三", "四", "五", "六", "七", "八", "九", "十")


def _has_different_branch_token(proc_tenderer: str, proc_cand: str) -> bool:
    """判断两个字符串是否因含不同的「分支机构特征词」而属于不同主体。

    用途：识别「苏州市公安局」vs「苏州市公安局交通警察支队」——虽然客户名
    出现在 tenderer 中，但"交通警察"是下级机构，不能算作同一客户。

    实现：用「在 tenderer 串中存在但候选串中不存在」的方向来检测特征词，
    因为典型场景是"更详细的 tenderer 比候选客户多出下级机构词"。
    """
    if not proc_tenderer or not proc_cand:
        return False
    # 方向 1：tenderer 比候选多出"分支机构特征词" → 不同主体
    for tok in _BRANCH_TOKENS:
        if tok in proc_tenderer and tok not in proc_cand:
            return True
        # 反向：候选比 tenderer 多出特征词（如"分行"在候选，tenderer 只有总行）
        if tok in proc_cand and tok not in proc_tenderer:
            return True
    return False


def _has_different_differentiator(proc_tenderer: str, proc_cand: str) -> bool:
    """判断两个字符串在「主体判别词」（序数词/地名/区域/字号）上是否不同。

    典型场景（应判为不同主体）：
      - 苏州大学附属第二医院 vs 苏州大学附属第一医院 （第二 vs 第一）
      - 苏州市虎丘山风景名胜区管理处 vs 苏州市盘门风景名胜区管理处 （虎丘山 vs 盘门）
      - 江苏苏州农村商业银行 vs 江苏东吴农村商业银行 （苏州 vs 东吴）

    典型场景（不应判为不同主体，子公司→集团 / 分公司→总部）：
      - 苏州市吴中产业投资集团 vs 苏州市产业投资集团 （吴中 是区域标识）

    实现：用 SequenceMatcher 找出「不同的子串」，如果双方各自含有某个
    判别词（且另一方没有）→ 视为不同主体。
    例外：当 candidate 出现在 tenderer 中（直接子串 或 去除"区域/子公司"
    标识后可作为子串）→ 这是「子公司→集团」场景，不算不同主体。
    """
    if not proc_tenderer or not proc_cand:
        return False
    if proc_tenderer == proc_cand:
        return False
    # 例外：candidate 出现在 tenderer 中
    #   - 直接子串：cp in proc
    #   - 去除 tenderer 中的判别词后能成为子串：tenderer 的差异片段全是判别词
    if proc_cand in proc_tenderer and len(proc_tenderer) > len(proc_cand):
        return False
    # 用 SequenceMatcher 找到「不匹配」的子串
    sm = SequenceMatcher(None, proc_tenderer, proc_cand)
    blocks = sm.get_matching_blocks()
    t_diff = []
    c_diff = []
    pos1 = pos2 = 0
    for blk in blocks:
        if blk.a > pos1:
            t_diff.append(proc_tenderer[pos1:blk.a])
        if blk.b > pos2:
            c_diff.append(proc_cand[pos2:blk.b])
        pos1 = blk.a + blk.size
        pos2 = blk.b + blk.size
    t_diff_text = "".join(t_diff)
    c_diff_text = "".join(c_diff)
    # 例外扩展：若 tenderer 相对 candidate 多出的部分全是「判别词」，
    # 这是「子公司→集团」场景（中间隔了个判别词），不算不同主体
    # 例：tenderer="苏州市吴中产业投资集团", candidate="苏州市产业投资集团"
    #     t_diff="吴中", c_diff=""  →  tenderer 多出 "吴中"，是判别词 → OK
    if not c_diff_text and t_diff_text:
        # tenderer 比 candidate 多，但 candidate 中没有"差异"内容
        # 检查 t_diff 全部由判别词组成
        rest = t_diff_text
        is_all_differentiator = True
        for tok in _DIFFERENTIATOR_TOKENS:
            while tok in rest:
                rest = rest.replace(tok, "", 1)
        # 也允许"分支词"（"吴中"既是判别词也是分支词）
        for tok in _BRANCH_TOKENS:
            while tok in rest:
                rest = rest.replace(tok, "", 1)
        if not rest.strip() and t_diff_text.strip():
            return False  # 全是判别词 → 子公司→集团
    # 1) 双方各自的"差异片段"中若各含一个判别词 → 不同主体
    for tok in _DIFFERENTIATOR_TOKENS:
        if tok in t_diff_text and tok not in c_diff_text and proc_tenderer.count(tok) > proc_cand.count(tok):
            return True
        if tok in c_diff_text and tok not in t_diff_text and proc_cand.count(tok) > proc_tenderer.count(tok):
            return True
    # 2) 短差异片段（1-2 字符）且双方各自含一个不同的中文数字序数 → 不同主体
    #    例：t_diff="二"  c_diff="一"  →  数字序数不同 → 不同主体
    if 1 <= len(t_diff_text) <= 2 and 1 <= len(c_diff_text) <= 2:
        t_digit = next((d for d in _CHINESE_DIGITS if d in t_diff_text), None)
        c_digit = next((d for d in _CHINESE_DIGITS if d in c_diff_text), None)
        if t_digit and c_digit and t_digit != c_digit:
            return True
    return False


def _substr_match(tenderer: str, cand: str) -> bool:
    """在剔除末尾停用词后，判断 cand 是否在 tenderer 中（子串匹配）。

    适用场景：招标人="苏州市发展和改革委员会投资处"，
             客户名="苏州市发展和改革委员会" → 应匹配
    """
    if not tenderer or not cand:
        return False
    t = preprocess(tenderer)
    c = preprocess(cand)
    if not t or not c:
        return False
    # 直接子串
    if c in t or t in c:
        return True
    return False


def load_kehu(path: str) -> tuple[list[str], dict[str, tuple[str, str]]]:
    """加载客户清单。
    返回：(客户名列表, {客户名: (客户经理, bu名称)} 映射)
    兼容旧格式（只有一列客户名）和新格式（含 area_name_t0/area_name_zj_t0）。
    """
    wb  = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws  = wb.active
    names = []
    info_map = {}
    # 读取表头，判断列结构
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    has_extra = header_row is not None and len(header_row) >= 3
    for row in ws.iter_rows(min_row=2, values_only=True):   # 跳过标题行
        v = row[0]
        if v and str(v).strip():
            name = str(v).strip()
            names.append(name)
            if has_extra:
                manager = str(row[1]).strip() if row[1] else ""
                bu      = str(row[2]).strip() if row[2] else ""
                info_map[name] = (manager, bu)
    wb.close()
    return names, info_map


def match_tenderer(tenderer: str, candidates: list[str], threshold: float) -> str:
    """
    多策略匹配：返回最相似的客户名；无相似客户则返回 "无"。

    匹配优先级（命中即返回）：
      A) 精确匹配：preprocess 后完全相等
      B) 客户名 ⊆ 招标人（招标人更长）：
         - 例：招标人=「苏州银行股份有限公司」/ 客户=「苏州银行股份有限公司唯亭西区支行」→ 不应匹配
         - 例：招标人=「苏州市发展和改革委员会投资处」/ 客户=「苏州市发展和改革委员会」→ 应匹配
         - 关键：若含"分支/上下级"特征词则不匹配（见 _has_different_branch_token）
      C) SequenceMatcher 相似度 ≥ threshold，取最高分；同样过滤分支误匹配
         - 新增：若含"主体判别词"差异（序数/地名/区域/字号）则不匹配
      D) 兜底：相似度 < threshold，返回 "无"

    设计要点：
      - 优先级 A 避免「苏州市立医院」被剥成「苏州市立」后误匹配
      - 优先级 B 要求「更短的一边 ≥ 4 字」避免单字/双字误匹配
      - 分支机构黑名单（_BRANCH_TOKENS）阻止「公安局 vs 交警支队」「分行 vs 支行」类错配
      - 主体判别词（_DIFFERENTIATOR_TOKENS）阻止「附属第一医院 vs 附属第二医院」
        「虎丘山 vs 盘门」「江苏苏州 vs 江苏东吴」类错配
    """
    if not tenderer or not tenderer.strip():
        return "无"

    orig = tenderer.strip()
    proc = preprocess(orig)
    if not proc:
        return "无"

    # 1) 精确匹配
    for cand in candidates:
        if proc == preprocess(cand):
            return cand

    # 2) 客户名 ⊆ 招标人（招标人更长），要求更短的一边长度 ≥ 4
    best_substr, best_substr_len = None, 0
    for cand in candidates:
        cp = preprocess(cand)
        if not cp or len(cp) < 4 or len(proc) < 4:
            continue
        # 客户名是 tenderer 子串（推荐场景：招标人含详细地址/部门，客户名是上级）
        if cp in proc:
            # 排除「分支/上下级」误匹配
            if _has_different_branch_token(proc, cp):
                continue
            # 排除「主体判别词」误匹配（序数/地名/区域/字号）
            if _has_different_differentiator(proc, cp):
                continue
            # 优先选「最长」的客户名（更精确）
            if len(cp) > best_substr_len:
                best_substr, best_substr_len = cand, len(cp)
    if best_substr:
        return best_substr

    # 3) 相似度匹配：取最高分，且不超过分支误匹配
    best, best_score = None, 0.0
    for cand in candidates:
        cp = preprocess(cand)
        if not cp:
            continue
        # 过滤分支/上下级
        if _has_different_branch_token(proc, cp):
            continue
        # 过滤主体判别词差异
        if _has_different_differentiator(proc, cp):
            continue
        score = SequenceMatcher(None, proc, cp).ratio()
        if score > best_score:
            best_score = score
            best = cand

    if best is not None and best_score >= threshold:
        return best
    return "无"


# ================================================================== #
#  AI 再核对：对低置信度结果请求大模型二次确认
# ================================================================== #
def ai_verify_match(tenderer: str, best_cand: str, score: float,
                    threshold: float) -> str:
    """
    在相似度刚好通过阈值、或存在歧义时，请求大模型再核对一次。
    仅在 _config 中启用了 ai_verify 时调用，避免默认增加 API 开销。
    返回：
        "YES:<客户名>" — AI 确认匹配
        "NO"           — AI 否定
        ""             — 调用失败/未启用，保持原结果
    """
    try:
        enabled = bool(_config.get("ai_verify", False))
    except Exception:
        enabled = False
    if not enabled:
        return ""
    # 只在边缘案例触发（分数靠近阈值），节省 API
    if score < threshold or score > threshold + 0.15:
        return ""
    if not tenderer or not best_cand:
        return ""

    try:
        prompt = (
            f"判断下面「招标人」与「客户名」是否指同一机构实体。\n"
            f"招标人：{tenderer}\n"
            f"客户名：{best_cand}\n"
            f"可以视为同一实体的差异（忽略）：\n"
            f"  - 括号内附加说明（如'（XX办公室）'）\n"
            f"  - 地址/校区/院区等后缀（'独墅湖校区'/'本部'等）\n"
            f"  - 简称差异（'委员会'与'委'、'有限公司'与'公司'、'管理局'与'局'）\n"
            f"必须视为不同实体的差异（不可忽略）：\n"
            f"  - 序数词不同（'第一医院' vs '第二医院'，'附一' vs '附二'）\n"
            f"  - 区域/地名不同（'虎丘山' vs '盘门'，'吴中' vs '姑苏'）\n"
            f"  - 企业字号不同（'江苏苏州' vs '江苏东吴'，'苏州银行' vs '东吴银行'）\n"
            f"  - 集团 vs 区域子公司（'苏州市产业投资集团' vs '苏州市吴中产业投资集团'）\n"
            f"只回答 YES 或 NO，不要其他文字。"
        )
        text = process.call_siliconflow(
            prompt,
            system_prompt="你是政府机构名称匹配助手，只回答 YES 或 NO。",
            max_retries=2, retry_delay=1.0,
        )
        if not text:
            return ""
        head = text.strip().upper().split()[0] if text.strip() else ""
        if head == "YES":
            return f"YES:{best_cand}"
        if head == "NO":
            return "NO"
        return ""
    except Exception:
        return ""


def ai_verify_tenderer(tenderer: str, candidates: list[str],
                        top_k: int = 5) -> str:
    """
    招标人再核对：把候选客户列表的前 top_k 提交给大模型，
    让其选出最可能的客户名（无则返回"无"）。
    用于 SequenceMatcher 评分的"模糊地带"或新增"无匹配"时的二次确认。
    返回：
        "YES:<客户名>" — AI 确认匹配
        "NO"           — AI 明确说无匹配
        ""             — 调用失败/未启用
    """
    try:
        enabled = bool(_config.get("ai_verify", False))
    except Exception:
        enabled = False
    if not enabled:
        return ""
    if not tenderer or not candidates:
        return ""

    # 仅取前 top_k（按 preprocess 后长度从大到小），节省 token
    cand_sorted = sorted(candidates, key=lambda x: len(preprocess(x)), reverse=True)[:top_k]
    cand_text = "\n".join(f"- {c}" for c in cand_sorted)

    try:
        prompt = (
            f"判断下面「招标人」是否与候选客户列表中某一个客户指同一机构。\n"
            f"招标人：{tenderer}\n"
            f"候选客户列表（{len(cand_sorted)} 个）：\n{cand_text}\n"
            f"注意：\n"
            f"1. 忽略括号内附加说明、地址/校区/院区等后缀\n"
            f"2. 忽略'委员会'与'委'、'有限公司'与'公司'等同类简称差异\n"
            f"3. 不要把上级机构与下级机构、分支机构当作同一客户\n"
            f"4. 不要把不同子公司/不同法人当作同一客户\n"
            f"请只输出客户列表中完全匹配的那一项（写完整名称），"
            f"或输出 NO（表示无匹配）。不要其他文字。"
        )
        text = process.call_siliconflow(
            prompt,
            system_prompt="你是政府机构名称匹配助手，只输出客户名或 NO。",
            max_retries=2, retry_delay=1.0,
        )
        if not text:
            return ""
        ans = text.strip()
        if ans.upper() == "NO":
            return "NO"
        # AI 可能返回的就是候选名之一；做精确校验
        for c in cand_sorted:
            if c in ans or ans in c or preprocess(c) == preprocess(ans):
                return f"YES:{c}"
        # 不在候选中视为"NO"
        return "NO"
    except Exception:
        return ""


def match_kehu(wb, ws, kehu_list: list[str], threshold: float = None,
               info_map: dict[str, tuple[str, str]] = None) -> tuple[int, int]:
    """
    对 Excel 工作表进行客户匹配，将结果写入 matched 列。
    若提供 info_map（客户名 → (客户经理, bu名称)），则匹配成功时同时写入
    "客户经理"和"bu名称"两列；并新增"方案跟进情况"列（默认空）。
    返回：(匹配数量, 未匹配数量)
    threshold 不指定时从 config 读取

    AI 二次核对策略（需在 config 中开启 ai_verify=True）：
      1) 边缘分数（threshold ~ threshold+0.15）：调用 ai_verify_match 单点确认
      2) 高相似度非子串匹配（ratio ≥ 0.80）：调 AI 防止「附属第一/第二医院」
         「虎丘山/盘门」「江苏苏州/江苏东吴」等结构相同但关键标识不同的误匹配
      3) 匹配结果 = "无"：调用 ai_verify_tenderer 把前 top_k 候选交给 AI 复核
    """
    if threshold is None:
        threshold = _get_threshold()
    total_rows  = ws.max_row
    matched     = 0
    not_matched = 0

    # 定位列索引
    col_map = {}
    for col_idx, (key, _, __) in enumerate(_COLUMNS, start=1):
        col_map[key] = col_idx
    matched_col   = col_map.get("matched")
    manager_col   = col_map.get("manager")
    bu_col        = col_map.get("bu")
    followup_col  = col_map.get("followup")

    if matched_col is None:
        return 0, 0

    print(f"\n开始匹配客户（阈值 ≥ {threshold}）...")

    for row_idx in range(2, total_rows + 1):          # 跳过标题行（第1行）
        tenderer = ws.cell(row=row_idx, column=2).value  # 第2列：招标人
        result   = match_tenderer(str(tenderer) if tenderer else "", kehu_list, threshold)

        # ---- AI 再核对：覆盖「边缘分数」和「高相似度非子串」两种可疑情况 ----
        if result != "无" and tenderer:
            try:
                t_proc = preprocess(str(tenderer).strip())
                c_proc = preprocess(result)
                if t_proc and c_proc:
                    score = SequenceMatcher(None, t_proc, c_proc).ratio()
                    cp = preprocess(result)
                    is_substring = (cp in t_proc) or (t_proc in cp)
                    # 可疑场景：
                    #   (a) 边缘分数（threshold ~ threshold+0.15）
                    #   (b) 高相似度（>= 0.80）但不是子串（典型误匹配）
                    suspicious = (
                        (threshold <= score <= threshold + 0.15) or
                        (score >= 0.80 and not is_substring)
                    )
                    if suspicious:
                        ai_result = ai_verify_match(
                            str(tenderer).strip(), result, score, threshold
                        )
                        if ai_result.startswith("YES:"):
                            result = ai_result.split(":", 1)[1]
                        elif ai_result == "NO":
                            result = "无"
            except Exception:
                pass

        # ---- 招标人再核对：原始匹配结果是"无"时，让 AI 在前 top_k 候选中挑 ----
        if result == "无" and tenderer:
            try:
                ai_result = ai_verify_tenderer(str(tenderer).strip(), kehu_list, top_k=5)
                if ai_result.startswith("YES:"):
                    result = ai_result.split(":", 1)[1]
                # AI 说"NO"则保持"无"
            except Exception:
                pass

        # 写入 matched 列
        cell = ws.cell(row=row_idx, column=matched_col)
        cell.value = result

        # 样式：匹配上→浅绿背景，未匹配→浅红背景
        if result == "无":
            cell.fill      = PatternFill("solid", fgColor="FFE6E6")
            cell.font      = Font(color="CC0000", size=9)
            not_matched += 1
            # 未匹配：客户经理/bu名称/方案跟进情况 留空
            if manager_col:
                ws.cell(row=row_idx, column=manager_col).value = ""
            if bu_col:
                ws.cell(row=row_idx, column=bu_col).value = ""
            if followup_col:
                ws.cell(row=row_idx, column=followup_col).value = ""
        else:
            cell.fill      = PatternFill("solid", fgColor="E6F4EA")
            cell.font      = Font(color="1A7A1A", size=9)
            matched += 1
            # 匹配成功：从 info_map 取客户经理/bu名称
            if info_map and result in info_map:
                manager, bu = info_map[result]
                if manager_col:
                    ws.cell(row=row_idx, column=manager_col).value = manager
                if bu_col:
                    ws.cell(row=row_idx, column=bu_col).value = bu
            else:
                if manager_col:
                    ws.cell(row=row_idx, column=manager_col).value = ""
                if bu_col:
                    ws.cell(row=row_idx, column=bu_col).value = ""
            # 方案跟进情况：默认空
            if followup_col:
                ws.cell(row=row_idx, column=followup_col).value = ""

        cell.alignment = Alignment(horizontal="left", vertical="center")
        title_short = str(tenderer or "")[:30]
        print(f"  [{row_idx-1}/{total_rows-1}] {title_short}... → {result}")

    return matched, not_matched


# ================================================================== #
#  第四阶段：导出 Excel
# ================================================================== #

# 列定义：(字段key, 列头中文, 列宽)
_COLUMNS = [
    ("title",      "公告标题",     45),
    ("tenderer",   "招标人",       20),
    ("date",       "发布日期",     13),
    ("result_due", "开标日期",     13),
    ("amount",     "预算金额",     14),
    ("is_it",      "信息化",        8),
    ("public_due", "公告期限",     14),
    ("source",     "数据来源",     10),
    ("matched",    "是否所属客户", 20),
    ("manager",    "客户经理",     14),
    ("bu",         "bu名称",       22),
    ("followup",   "方案跟进情况", 16),
    ("url",        "公告链接",     50),
]

def export_excel(records: list, xlsx_file: str = "extracted_info.xlsx") -> None:
    """
    将结果列表导出为带格式的 Excel 文件。
    标题行冻结，超链接列自动生成可点击链接。
    如果客户清单文件存在，则自动进行客户匹配。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "招标公告"

    # ----- 样式定义 -----
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    link_font    = Font(color="0563C1", underline="single", size=9)
    normal_font  = Font(size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_side    = Side(style="thin", color="BFBFBF")
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)

    # ----- 写表头 -----
    headers = [col[1] for col in _COLUMNS]
    ws.append(headers)
    for col_idx, (_, header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ----- 写数据 -----
    url_col_idx = next(
        (i + 1 for i, (k, _, __) in enumerate(_COLUMNS) if k == "url"), None
    )
    for row_idx, record in enumerate(records, start=2):
        for col_idx, (key, _, __) in enumerate(_COLUMNS, start=1):
            value = record.get(key) or ""
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.alignment = left_align
            cell.font      = normal_font

            # url 列设置超链接
            if col_idx == url_col_idx and value:
                cell.hyperlink = value
                cell.font      = link_font

        # 交替行背景
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="EBF3FB")
            for col_idx in range(1, len(_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # ----- 冻结首行 -----
    ws.freeze_panes = "A2"

    # ----- 行高 -----
    ws.row_dimensions[1].height = 20
    for row_idx in range(2, len(records) + 2):
        ws.row_dimensions[row_idx].height = 32

    # ----- 客户匹配（如果客户清单存在）-----
    if os.path.exists(KEHU_XLSX):
        print(f"  检测到客户清单 {KEHU_XLSX}，开始匹配...")
        kehu_list, info_map = load_kehu(KEHU_XLSX)
        print(f"  已加载 {len(kehu_list)} 个客户名")
        matched, not_matched = match_kehu(wb, ws, kehu_list, info_map=info_map)
        print(f"  匹配完成：{matched} 条匹配，{not_matched} 条未匹配")
    else:
        print(f"  未找到客户清单 {KEHU_XLSX}，跳过客户匹配")

    # ----- 保存文件（处理覆盖问题）-----
    if os.path.exists(xlsx_file):
        try:
            os.remove(xlsx_file)
            print(f"  已删除旧文件 {xlsx_file}")
        except PermissionError:
            print(f"\n  ❌ 错误：无法删除 {xlsx_file}")
            print(f"  该文件可能正被其他程序（如 Excel）打开，请关闭后重试。")
            print(f"  临时文件已保存至 {xlsx_file}.tmp")
            wb.save(f"{xlsx_file}.tmp")
            return
        except Exception as e:
            print(f"\n  ❌ 删除文件时出错：{e}")
            print(f"  临时文件已保存至 {xlsx_file}.tmp")
            wb.save(f"{xlsx_file}.tmp")
            return

    wb.save(xlsx_file)
    print(f"  Excel 已保存至 {xlsx_file}（共 {len(records)} 行）")


def run_phase2(input_file: str, output_xlsx: str = None) -> list:
    """
    运行 Phase 2，处理输入文件，提取字段并导出 Excel 结果。
    只生成 Excel（不生成 JSON），保存到 ./data/phase2/ 目录下（带时间戳）。
    返回：最终结果列表
    """
    print("\n" + "█"*50)
    print(f"  Phase 2: 处理采集到的信息")
    print(f"  输入文件: {input_file}")
    print("█"*50)

    # 确保输出目录存在
    os.makedirs(PHASE2_DIR, exist_ok=True)

    # 读取输入文件
    if not os.path.exists(input_file):
        print(f"错误：输入文件 {input_file} 不存在！")
        sys.exit(1)

    with open(input_file, "r", encoding="utf-8") as f:
        collected = json.load(f)

    print(f"已加载 {len(collected)} 条采集数据")

    if not collected:
        print("没有数据需要处理，Phase 2 结束。")
        return []

    # AI 字段提取
    final = extract_fields(collected)

    # 统一命名：phase2_extracted_YYYYMMDD_HHMMSS.xlsx，保存到 ./data/phase2/
    if output_xlsx is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = os.path.join(PHASE2_DIR, f"phase2_extracted_{ts}.xlsx")

    # 只导出 Excel（不生成 JSON）
    export_excel(final, output_xlsx)

    print(f"\n{'█'*50}")
    print(f"  Phase 2 完成！共处理 {len(final)} 条公告")
    print(f"  Excel → {output_xlsx}")
    print(f"{'█'*50}\n")

    return final


# ================================================================== #
#  完整流水线（Phase 1 + Phase 2）
# ================================================================== #

def run_full_pipeline() -> list:
    """
    运行完整流程：Phase 1 采集 → Phase 2 处理。
    返回：最终结果列表
    """
    # Phase 1
    collected, phase1_output = run_phase1()
    
    if not collected:
        return []
    
    # Phase 2
    final = run_phase2(phase1_output)
    
    return final


# ================================================================== #
#  增量处理：在 Phase 2 前去掉已提取过的公告
# ================================================================== #

def run_phase2_incremental(input_file: str, reference_files: list = None,
                           use_all_history: bool = False,
                           output_xlsx: str = None) -> list:
    """
    增量 Phase 2：先去掉已处理过的公告，再进行模型提取。
    input_file:       Phase 1 输出的 JSON 文件路径
    reference_files:  作为「已处理」参考的 phase2 Excel 文件名列表
                      （仅当 use_all_history=False 时有效）
    use_all_history:  为 True 时读取全部 phase2 Excel 文件作为参考
    output_xlsx:      输出 Excel 路径，不指定则自动生成
    返回：最终结果列表
    """
    print("\n" + "█"*50)
    print(f"  Phase 2 增量处理")
    print(f"  输入文件: {input_file}")
    print("█"*50)

    os.makedirs(PHASE2_DIR, exist_ok=True)

    if not os.path.exists(input_file):
        print(f"错误：输入文件 {input_file} 不存在！")
        return []

    with open(input_file, "r", encoding="utf-8") as f:
        collected = json.load(f)

    print(f"已加载 {len(collected)} 条采集数据")

    # 收集已处理记录
    if use_all_history:
        processed_keys = _history.collect_all_processed_keys()
    elif reference_files:
        processed_keys = _history.collect_processed_keys(reference_files)
    else:
        processed_keys = set()

    # 过滤掉已处理的公告
    if processed_keys:
        unprocessed, skipped = _history.filter_unprocessed(collected, processed_keys)
        print(f"增量去重：{len(collected)} 条 → 去掉已处理 {skipped} 条 → 待处理 {len(unprocessed)} 条")
        collected = unprocessed
    else:
        print("无历史记录，全量处理")

    if not collected:
        print("没有新数据需要处理，Phase 2 结束。")
        return []

    # AI 字段提取
    final = extract_fields(collected)

    if output_xlsx is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = os.path.join(PHASE2_DIR, f"phase2_extracted_{ts}.xlsx")

    export_excel(final, output_xlsx)

    print(f"\n{'█'*50}")
    print(f"  Phase 2 增量处理完成！共处理 {len(final)} 条新公告")
    print(f"  Excel → {output_xlsx}")
    print(f"{'█'*50}\n")

    return final


def run_full_incremental() -> list:
    """
    增量完整流程：Phase 1 采集 → 增量 Phase 2（去掉全部已处理记录）。
    用于每日/每周定时任务。
    返回：最终结果列表
    """
    collected, phase1_output = run_phase1()

    if not collected:
        return []

    final = run_phase2_incremental(phase1_output, use_all_history=True)
    return final


# ================================================================== #
#  命令行入口
# ================================================================== #

def main():
    parser = argparse.ArgumentParser(
        description="招标公告采集与处理流水线 (v3.0)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  # 运行完整流程（Phase 1 + Phase 2）
  python pipeline_v3.py
  
  # 仅运行 Phase 1（采集）
  python pipeline_v3.py --phase 1
  
  # 仅运行 Phase 2（处理），需指定输入文件
  python pipeline_v3.py --phase 2 --input data/phase1/phase1_collected_20260414_152030.json
  
  # 指定输出文件名
  python pipeline_v3.py --phase 1 --output data/phase1/collected_data.json

输出目录：
  Phase 1 最终文件 → ./data/phase1/phase1_collected_YYYYMMDD_HHMMSS.json
  Phase 2 最终文件 → ./data/phase2/phase2_extracted_YYYYMMDD_HHMMSS.xlsx
        """)
    
    parser.add_argument(
        "--phase", "-p",
        type=int,
        choices=[1, 2],
        help="指定运行阶段：1 (采集) 或 2 (处理)。如果不指定，则运行完整流程。"
    )
    
    parser.add_argument(
        "--input", "-i",
        type=str,
        help="Phase 2 的输入文件路径（JSON 格式）。仅在 --phase 2 时必需。"
    )
    
    parser.add_argument(
        "--output", "-o",
        type=str,
        help="Phase 1 的输出文件名。如果不指定，则自动生成带时间戳的文件名。"
    )
    
    parser.add_argument(
        "--xlsx-out",
        type=str,
        help="Phase 2 的 Excel 输出文件名（覆盖默认时间戳命名）。"
    )
    
    args = parser.parse_args()
    
    if args.phase == 1:
        # 仅运行 Phase 1
        run_phase1(args.output)
    
    elif args.phase == 2:
        # 仅运行 Phase 2
        if not args.input:
            print("错误：运行 Phase 2 必须指定 --input 参数！")
            sys.exit(1)
        run_phase2(args.input, args.xlsx_out)
    
    else:
        # 未指定 phase，运行完整流程
        run_full_pipeline()


if __name__ == "__main__":
    main()
